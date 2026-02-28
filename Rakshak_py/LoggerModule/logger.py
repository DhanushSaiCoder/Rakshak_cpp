import threading
import queue
import struct
from openpyxl import Workbook
from .global_config import settings
from .storage import SystemStorage
from .file_manager import FileManager

class Logger:
    def __init__(self, file_type = None, compress = None, file_keyword = None):
        try:
            self.file_manager = None
            self._running = False
            self.file_type = file_type or settings._DEFAULT_FILE_TYPE
            self.schema = None
            self._enabled = True
            self.file_keyword = file_keyword or None

            self.q = queue.Queue(maxsize=settings._QUEUE_SIZE)
            self.headers_blob = None
            self.dropped_count = 0

            self._worker = None
            self.do_compress = compress if compress is not None else settings._DEFAULT_COMPRESS
            self.initialize()
            

        except Exception as e:
            print(f"Exception in init: {e}")

    # =========================== INITILIZER ========================
    def initialize(self):
        try:
            if not SystemStorage().checking():
                self._enabled = False
                self._running = False
                self.file_manager = None
                return 


            self.file_manager = FileManager(file_type=self.file_type, compress=self.do_compress, file_keyword=self.file_keyword)

        except Exception as e:
            print(f"Exception in initilizer: {e}")
            self._enabled = False

    # =========================== START ========================
    def start(self):
        try:
            if not self._enabled or self.file_manager is None:
                return   

            self._running = True

            match self.file_type:
                case "csv":
                    self._worker = threading.Thread(target=self.csv_worker, daemon=True)
                case "bin":
                    self._worker = threading.Thread(target=self.bin_worker, daemon=True)
                case "tlv":
                    self._worker = threading.Thread(target=self.tlv_worker, daemon=True)
                case "xlsx":
                    self._worker = threading.Thread(target=self.xlsx_worker, daemon=True)
                case _:
                    return

            self._worker.start()

        except Exception as e:
            print(f"Exception in start: {e}")
            self._enabled = False

    # =========================== HEADER WRITER ========================
    def headers(self, *headers):
        try:
            if not self._enabled:
                return

            if not headers:
                raise ValueError("Headers cannot be none")

            TYPE_NULL = 0
            TYPE_BOOL = 1
            TYPE_INT = 2
            TYPE_FLOAT = 3
            TYPE_STRING = 4

            clean_schema = []
            column_types = []

            for h in headers:
                if ":" in h:
                    name, type_str = h.split(":", 1)
                    name = name.strip()
                    type_str = type_str.strip().lower()
                    
                    clean_schema.append(name)
                    
                    if "int" in type_str: column_types.append(TYPE_INT)
                    elif "float" in type_str: column_types.append(TYPE_FLOAT)
                    elif "bool" in type_str: column_types.append(TYPE_BOOL)
                    elif "str" in type_str: column_types.append(TYPE_STRING)
                    else: column_types.append(TYPE_FLOAT)
                else:
                    clean_schema.append(h)
                    column_types.append(TYPE_FLOAT)

            self.schema = tuple(clean_schema)

            
            if self.file_type == "bin":
                if not column_types:
                    raise ValueError("Bin format needs column types")
                buf = bytearray()
                buf += b"LOG1"
                buf += (1).to_bytes(1, "little")
                buf += len(self.schema).to_bytes(2, "little")
                
                for name, type_id in zip(self.schema, column_types):
                    b_name = name.encode("utf-8")
                    buf += len(b_name).to_bytes(2, "little")
                    buf += b_name
                    buf += type_id.to_bytes(1, "little") 
                
                self.headers_blob = bytes(buf)

            elif self.file_type == "tlv":
                buf = bytearray()
                buf += b"TLV1"
                buf += (1).to_bytes(1, "little")
                buf += len(self.schema).to_bytes(1, "little")
                FIELD_DEF = 0x01
                
                for name in self.schema:
                    b = name.encode("utf-8")
                    buf += FIELD_DEF.to_bytes(1, "little")
                    buf += len(b).to_bytes(2, "little")
                    buf += b
                
                self.headers_blob = bytes(buf)

            elif self.file_type == "csv":
                self.headers_blob = (",".join(self.schema) + "\n").encode("utf-8")

            elif self.file_type == "xlsx":
                self.headers_blob = self.schema

            else:
                raise ValueError(f"Unsupported file type: {self.file_type}")

        except Exception as e:
            print(f"Exception in header: {e}")
            self._enabled = False

    # =========================== PUBLISHER ========================
    def publish(self, values=None, encode=None):
        
        if not self._running or not self._enabled:
            return

        try:
            if values is None:
                raise ValueError("Values cannot be None")

            record = None
            if encode == None:
                encode = settings._ENCODER

            # =======================
            # 1. NORMALIZE INPUT
            # =======================
            if isinstance(values, dict):
                if not self.schema:
                    raise RuntimeError(
                        "Schema not set. Call headers() before using dict input."
                    )

                # -------- INLINE FLATTEN (dict + list) --------
                flat = {}
                stack = [(values, "")]

                while stack:
                    current, prefix = stack.pop()

                    if isinstance(current, dict):
                        for k, v in current.items():
                            key = f"{prefix}.{k}" if prefix else k
                            stack.append((v, key))

                    elif isinstance(current, list):
                        for i, v in enumerate(current):
                            key = f"{prefix}[{i}]"
                            stack.append((v, key))

                    else:
                        flat[prefix] = current


                processed_values = [flat.get(k) for k in self.schema]

            elif isinstance(values, (list, tuple)):
                processed_values = list(values)

            elif isinstance(values, (bytes, bytearray)):
                processed_values = None  

            else:
                raise TypeError(
                    f"Unsupported input type: {type(values)}"
                )

            # =======================
            # 2. FORMAT HANDLING
            # =======================
            if self.file_type in ("bin", "tlv"):

                # ---- RAW BINARY PATH ----
                if not encode:
                    if not isinstance(values, (bytes, bytearray)):
                        raise TypeError(
                            "encode=False requires raw bytes input or make it encode = True"
                        )
                    record = values

                # ---- STRUCTURED → BINARY ----
                else:
                    if processed_values is None:
                        raise TypeError(
                            "encode=True requires structured input (list/dict) or make it encode = False"
                        )

                    if self.file_type == "bin":
                        record = self._encode_record_bin(processed_values)
                    else:
                        record = self._encode_record_tlvbin(processed_values)

            elif self.file_type in ("csv", "xlsx"):
                if encode:
                    raise TypeError(
                        "Encoder = True CSV/XLSX doesn't need make it Encoder =  False"
                    )
                
                if processed_values is None:
                    raise TypeError(
                        "CSV/XLSX do not accept raw binary input"
                    )
                record = processed_values

            else:
                raise ValueError(f"Unsupported file type: {self.file_type}")

            # =======================
            # 3. QUEUE
            # =======================
            try:
                self.q.put(record, timeout=0.001)
            except queue.Full:
                self.dropped_count += 1

        except Exception as e:
            self._enabled = False
            print(f"Exception in publish: {e}")

    # =========================== STOP ========================
    def stop(self):
        try:
            self._enabled = False
            self._running = False
            
            if self._worker and self._worker.is_alive():
                # Optional: Add a timeout to the join to prevent infinite hangs
                # self.q.join() 
                self._worker.join(timeout=2.0)
        except Exception as e:
            print(f"Exception in stop: {e}")

    # =========================== ENCODERS ========================
    def _encode_record_bin(self, values):
            try:
                payload = bytearray()

                for value in values:
                    if value is None:
                        payload += struct.pack("<q", -9223372036854775808)
                        
                    elif isinstance(value, bool):
                        payload += struct.pack("<q", 1 if value else 0)
                    elif isinstance(value, int):
                        payload += struct.pack("<q", value)
                    elif isinstance(value, float):
                        payload += struct.pack("<d", value)
                    elif isinstance(value, str):
                        payload += value.encode("utf-8") + b"\x00"
                    else:
                        raise TypeError(f"Unsupported type for BIN: {type(value)}")
                return bytes(payload)
            except Exception as e:
                print(f"Exception in _encode_record_bin: {e}")  



    def _encode_record_tlvbin(self, values):
        try:
            if not self.schema:
                raise RuntimeError("Schema not set. Call headers() first.")

            if len(values) != len(self.schema):
                raise ValueError("Record does not match schema length")

            TYPE_BOOL   = 1
            TYPE_INT    = 2
            TYPE_FLOAT  = 3
            TYPE_STRING = 4
            TYPE_BYTES  = 5
            TYPE_NONE   = 6

            buf = bytearray()

            for value in values:


                if isinstance(value, (dict, list, tuple)):
                    raise TypeError(
                        "TLV encoder received non-leaf value. "
                        "Normalization error (dict/list should be flattened)."
                    )

                # ---------- NONE ----------
                if value is None:
                    buf += TYPE_NONE.to_bytes(1, "little")
                    buf += (0).to_bytes(2, "little")

                # ---------- BOOL ----------
                elif isinstance(value, bool):
                    buf += TYPE_BOOL.to_bytes(1, "little")
                    buf += (1).to_bytes(2, "little")
                    buf += b"\x01" if value else b"\x00"

                # ---------- INT ----------
                elif isinstance(value, int):
                    data = struct.pack("<q", value)  
                    buf += TYPE_INT.to_bytes(1, "little")
                    buf += len(data).to_bytes(2, "little")
                    buf += data

                # ---------- FLOAT ----------
                elif isinstance(value, float):
                    data = struct.pack("<d", value)  
                    buf += TYPE_FLOAT.to_bytes(1, "little")
                    buf += len(data).to_bytes(2, "little")
                    buf += data

                # ---------- STRING ----------
                elif isinstance(value, str):
                    data = value.encode("utf-8")
                    if len(data) > 0xFFFF:
                        raise ValueError("TLV string too large")
                    buf += TYPE_STRING.to_bytes(1, "little")
                    buf += len(data).to_bytes(2, "little")
                    buf += data

                # ---------- BYTES ----------
                elif isinstance(value, (bytes, bytearray)):
                    if len(value) > 0xFFFF:
                        raise ValueError("TLV bytes too large")
                    buf += TYPE_BYTES.to_bytes(1, "little")
                    buf += len(value).to_bytes(2, "little")
                    buf += value

                else:
                    raise TypeError(f"Unsupported type: {type(value)}")

            record = bytearray()
            record += len(buf).to_bytes(2, "little")
            record += buf

            return bytes(record)

        except Exception as e:
            raise RuntimeError(f"Exception in TLV encoder: {e}")



        
    # =========================== BIN WORKER ========================
    def bin_worker(self):
        try:
            max_bytes = settings._MAX_FILE_SIZE_MB * 1024 * 1024
            current_size = 0
            
            f = open(self.file_manager.current_file, "ab")

            f.write(self.headers_blob)
            f.flush()

            current_size = len(self.headers_blob)

            try:
                while self._running or not self.q.empty():
                    try:
                        record = self.q.get(timeout=0.1)
                    except queue.Empty:
                        continue

                    size = len(record)

                    if current_size + size >= max_bytes:
                        f.close()

                        self.file_manager.current_file = self.file_manager._new_log_file()
                        f = open(self.file_manager.current_file, "ab")
                        current_size = 0

                        if self.headers_blob:
                            f.write(self.headers_blob)
                            f.flush()
                            current_size += len(self.headers_blob)

                        self.file_manager.compress_logs()


                    f.write(record)
                    current_size += size

                    self.q.task_done()

            finally:
                f.close()
        except Exception as e:
            self._running = False
            print(f"Exception in Bin Worker: {e}")

    # =========================== TLV BIN WORKER ========================
    def tlv_worker(self):
        try:
            max_bytes = settings._MAX_FILE_SIZE_MB * 1024 * 1024
            current_size = 0
            f = open(self.file_manager.current_file, "ab")

            if self.headers_blob:
                f.write(self.headers_blob)
                f.flush()
                current_size = len(self.headers_blob)

            while self._running or not self.q.empty():
                try:
                    record = self.q.get(timeout=0.1)
                except queue.Empty:
                    continue

                size = len(record)
                if current_size + size >= max_bytes:
                    f.close()
                    self.file_manager.current_file = self.file_manager._new_log_file()
                    f = open(self.file_manager.current_file, "ab")
                    current_size = 0

                    if self.headers_blob:
                        f.write(self.headers_blob)
                        f.flush()
                        current_size += len(self.headers_blob)
                    
                    self.file_manager.compress_logs()

                f.write(record)
                current_size += size
                self.q.task_done()
        except Exception as e:
            self._running = False
            print(f"Exception in TLV worker: {e}")
        finally:
            f.close()


    # =========================== CSV WORKER ========================
    def csv_worker(self):
        try:
            max_bytes = settings._MAX_FILE_SIZE_MB * 1024 * 1024
            current_size = 0

            f = open(self.file_manager.current_file, "ab")

            if self.headers_blob:
                f.write(self.headers_blob)
                f.flush()
                current_size += len(self.headers_blob)

            try:
                while self._running or not self.q.empty():
                    try:
                        record = self.q.get(timeout=0.1)
                    except queue.Empty:
                        continue

                    line = ",".join(map(str, record)) + "\n"

                    encoded = line.encode("utf-8")
                    size = len(encoded)

                    if current_size + size >= max_bytes:
                        f.close()

                        self.file_manager.current_file = self.file_manager._new_log_file()
                        f = open(self.file_manager.current_file, "ab")
                        current_size = 0

                        if self.headers_blob:
                            f.write(self.headers_blob)
                            f.flush()
                            current_size += len(self.headers_blob)

                        self.file_manager.compress_logs()

                    f.write(encoded)
                    current_size += size
                    self.q.task_done()

            finally:
                f.close()
        except Exception as e:
            self._running = False
            print(f"Exception in CSV worker: {e}")


    def xlsx_worker(self):
        try:
            # Excel hard limit ≈ 1,048,576
            MAX_ROWS = settings._XLSX_MAX_ROWS
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="log")
            row_count = 0

            def prepare_new_sheet(workbook):
                sheet = workbook.create_sheet(title="log")
                count = 0
                if self.schema:
                    sheet.append(list(self.schema))
                    count = 1
                return sheet, count

            if self.schema:
                ws.append(list(self.schema))
                row_count = 1

            try:
                while self._running or not self.q.empty():
                    try:
                        record = self.q.get(timeout=0.5)
                    except queue.Empty:
                        continue

                    try:
                        ws.append(list(record))
                        row_count += 1
                    finally:

                        self.q.task_done()

                    
                    if row_count >= MAX_ROWS:
                        wb.save(self.file_manager.current_file)
                        

                        # Setup new workbook
                        self.file_manager.current_file = self.file_manager._new_log_file()
                        wb = Workbook(write_only=True)
                        ws, row_count = prepare_new_sheet(wb)

                        self.file_manager.compress_logs()


            except Exception as e:
                print(f"Worker error: {e}")
            finally:
                try:
                    wb.save(self.file_manager.current_file)
                    wb.close()
                except Exception:
                    pass

        except Exception as e:
            self._running = False
            print(f"Exception in XLSX Worker: {e}")
